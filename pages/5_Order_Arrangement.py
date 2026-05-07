# pages/5_Order_Arrangement.py

import streamlit as st
import pandas as pd
from supabase import create_client, Client
import datetime
import random
import io
import json
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill

# --- 1. 配置、CSS样式 ---
SUPABASE_URL = st.secrets["supabase_url"]
SUPABASE_KEY = st.secrets["supabase_key"]
supabase: Client = create_client(SUPABASE_URL, SUPABASE_KEY)

st.set_page_config(layout="wide")
st.title("📦 订单规划")

st.markdown("""
<style>
.status-tag { display: inline-block; padding: 2px 10px; margin: 2px; font-size: 0.9em; font-weight: 500; border-radius: 12px; border: 1px solid; }
.status-completed { color: #2a9d8f; border-color: #2a9d8f; background-color: #f0fafa;}
.status-inprogress { color: #e76f51; border-color: #e76f51; background-color: #fff8f5;}
.status-notstarted { color: #6c757d; border-color: #6c757d; background-color: #f8f9fa;}
.stNumberInput > div > div > input { text-align: center; }
</style>
""", unsafe_allow_html=True)

# --- 2. 数据加载函数 ---
@st.cache_data(ttl=30)
def load_data():
    cart_query = "id, client_websites(id, name, projects(id, project_name)), vendor_offers(price, vendors(name), referring_domains(domain_name))"
    urls_response = supabase.table('website_urls').select('client_website_id, url').execute()
    anchors_response = supabase.table('anchor_texts').select('project_id, text, search_volume').execute()
    cart_response = supabase.table('shopping_cart').select(cart_query).execute()
    
    # ⭐️ 核心改动：现在我们还需要 client_websites 的 name 来显示哪个模板属于哪个网站
    templates_response = supabase.table('planning_templates').select('id, website_id, combinations_json, client_websites(name)').execute()
    
    cart_df = pd.json_normalize(cart_response.data, sep='.') if cart_response.data else pd.DataFrame()
    urls_df = pd.DataFrame(urls_response.data) if urls_response.data else pd.DataFrame()
    
    if anchors_response.data:
        anchors_df = pd.DataFrame(anchors_response.data)
        anchors_df['search_volume'] = pd.to_numeric(anchors_df['search_volume'], errors='coerce').fillna(0)
    else:
        anchors_df = pd.DataFrame(columns=['project_id', 'text', 'search_volume'])
        
    templates_data = templates_response.data if templates_response.data else []
    
    return cart_df, urls_df, anchors_df, templates_data

# --- 3. 主逻辑 ---
cart_df, urls_df, anchors_df, templates_data = load_data()

if 'preview_df' not in st.session_state: st.session_state.preview_df = None
if 'combinations' not in st.session_state: st.session_state.combinations = {}
if 'selected_vendors_for_download' not in st.session_state: st.session_state.selected_vendors_for_download = []


if cart_df.empty:
    st.success("购物车是空的。")
else:
    website_counts = cart_df['client_websites.name'].value_counts().to_dict()
    website_info_map = cart_df.drop_duplicates(subset=['client_websites.name']).set_index('client_websites.name')[
        ['client_websites.id', 'client_websites.projects.id']
    ].to_dict('index')
    url_map = urls_df.groupby('client_website_id')['url'].apply(list).to_dict()

    saved_templates = {item['website_id']: item['combinations_json'] for item in templates_data}

    st.header("🗓️ 1. 周期设置")
    g_col1, g_col2 = st.columns(2)
    start_date = g_col1.date_input("开始日期", datetime.date.today())
    end_date = g_col2.date_input("结束日期", datetime.date.today() + datetime.timedelta(days=30))
    st.markdown("---")
    
    st.header("✍️ 2. 组合策略")
    
    # ⭐️ 核心改动：添加一个可展开的管理区域
    with st.expander("⚙️ 管理已保存的策略"):
        if not templates_data:
            st.info("目前没有已保存的策略可供管理。")
        else:
            st.warning("在这里，您可以永久删除不再需要的已存策略。")
            
            # 将模板数据转换为 DataFrame 以便使用 st.data_editor
            templates_df = pd.json_normalize(templates_data, sep='_')
            templates_df = templates_df.rename(columns={'client_websites_name': '网站名称'})
            templates_df.insert(0, "选择删除", False)
            
            edited_templates_df = st.data_editor(
                templates_df[['选择删除', '网站名称']], # 只显示这两列
                key="delete_templates_editor",
                disabled=['网站名称'],
                hide_index=True,
                width='stretch'
            )
            
            templates_to_delete = edited_templates_df[edited_templates_df['选择删除']]
            
            if not templates_to_delete.empty:
                if st.button(f"🗑️ 删除所选的 {len(templates_to_delete)} 个策略", type="primary"):
                    # 获取原始 DataFrame 中对应要删除的行的 ID
                    ids_to_delete = templates_df.loc[templates_to_delete.index, 'id'].tolist()
                    with st.spinner("正在删除..."):
                        try:
                            supabase.table('planning_templates').delete().in_('id', ids_to_delete).execute()
                            st.success("成功删除所选策略！")
                            # 清理缓存和 session_state 以反映变化
                            st.cache_data.clear()
                            # 如果被删除的策略正好是当前加载的，也一并清理
                            for index, row in templates_to_delete.iterrows():
                                site_name_to_clear = row['网站名称']
                                if site_name_to_clear in st.session_state.combinations:
                                    del st.session_state.combinations[site_name_to_clear]
                            st.rerun()
                        except Exception as e:
                            st.error(f"删除失败: {e}")

    cols_per_row = 6
    all_cols = st.columns(cols_per_row)
    col_idx = 0
    for name, count in website_counts.items():
        combos = st.session_state.combinations.get(name, [])
        total_count = sum(c.get('count', 0) for c in combos)
        if total_count == 0: status, css = "未开始", "status-notstarted"
        elif total_count == count: status, css = "已完成", "status-completed"
        else: status, css = "进行中", "status-inprogress"
        all_cols[col_idx % cols_per_row].markdown(f'<div class="status-tag {css}">{name} ({status})</div>', unsafe_allow_html=True)
        col_idx += 1
    st.markdown("---")

    selected_website = st.selectbox("选择要规划的网站", options=["-- 请选择 --"] + list(website_counts.keys()))

    if selected_website != "-- 请选择 --":
        count = website_counts[selected_website]
        info = website_info_map.get(selected_website, {})
        website_id = info.get('client_websites.id')
        project_id = info.get('client_websites.projects.id')
        
        left_col, right_col = st.columns([2, 1])

        with left_col:
            if selected_website not in st.session_state.combinations:
                if website_id in saved_templates:
                    st.session_state.combinations[selected_website] = saved_templates[website_id]
                    st.toast(f"✅ 已成功加载 '{selected_website}' 的保存策略！")
                else:
                    st.session_state.combinations[selected_website] = [{'a1': '-- 选择锚文本 --', 'u1': '-- 选择URL --', 'a2': '-- 选择锚文本 --', 'u2': '-- 选择URL --', 'count': 0} for _ in range(3)]

            st.info(f"正在为网站 **{selected_website}** 规划 **{count}** 个链接。")
            
            combos = st.session_state.combinations[selected_website]
            available_urls = ["-- 选择URL --"] + url_map.get(website_id, [])
            project_anchors_df = anchors_df[anchors_df['project_id'] == project_id] if project_id is not None else pd.DataFrame(columns=['text', 'search_volume'])
            anchor_options = ["-- 选择锚文本 --"] + sorted(project_anchors_df['text'].tolist())

            header_cols = st.columns([6, 6, 2, 1])
            header_cols[0].markdown("**锚文本 (Anchor Text)**")
            header_cols[1].markdown("**目标链接 (URL)**")
            header_cols[2].markdown("<div style='text-align:center'>数量</div>", unsafe_allow_html=True)
            st.markdown("---")

            total_combo_count = 0
            indices_to_remove = []
            for i, combo in enumerate(combos):
                row1_cols = st.columns([6, 6, 2, 1])
                with row1_cols[0]: combo['a1'] = st.selectbox("Anchor_1", anchor_options, index=anchor_options.index(combo['a1']) if combo['a1'] in anchor_options else 0, key=f"a1_{selected_website}_{i}", label_visibility="collapsed")
                with row1_cols[1]: combo['u1'] = st.selectbox("Url_1", available_urls, index=available_urls.index(combo['u1']) if combo['u1'] in available_urls else 0, key=f"u1_{selected_website}_{i}", label_visibility="collapsed")
                with row1_cols[2]: combo['count'] = st.number_input("数量", min_value=0, value=combo['count'], key=f"count_{selected_website}_{i}", label_visibility="collapsed")
                with row1_cols[3]:
                    if st.button("🗑️", key=f"del_{selected_website}_{i}", help="删除此组合"): indices_to_remove.append(i)
                
                row2_cols = st.columns([6, 6, 3])
                with row2_cols[0]: combo['a2'] = st.selectbox("Anchor_2", anchor_options, index=anchor_options.index(combo['a2']) if combo['a2'] in anchor_options else 0, key=f"a2_{selected_website}_{i}", label_visibility="collapsed")
                with row2_cols[1]: combo['u2'] = st.selectbox("Url_2", available_urls, index=available_urls.index(combo['u2']) if combo['u2'] in available_urls else 0, key=f"u2_{selected_website}_{i}", label_visibility="collapsed")
                with row2_cols[2]:
                    percentage = (combo['count'] / count * 100) if count > 0 else 0
                    st.markdown(f"<div style='text-align: center; font-size: 0.9em; color: #888; padding-top: 8px;'>({percentage:.1f}%)</div>", unsafe_allow_html=True)
                
                st.markdown("<div style='margin-bottom: 10px;'></div>", unsafe_allow_html=True)
                total_combo_count += combo['count']

            if indices_to_remove:
                for index in sorted(indices_to_remove, reverse=True): combos.pop(index)
                st.rerun()

            btn_cols = st.columns(2)
            with btn_cols[0]:
                if st.button(f"➕ 添加一个新组合", width='stretch'):
                    combos.append({'a1': '-- 选择锚文本 --', 'u1': '-- 选择URL --', 'a2': '-- 选择锚文本 --', 'u2': '-- 选择URL --', 'count': 0})
                    st.rerun()
            
            with btn_cols[1]:
                if st.button("💾 保存当前策略", type="primary", width='stretch'):
                    with st.spinner("正在保存策略到数据库..."):
                        try:
                            supabase.table('planning_templates').upsert({
                                'website_id': website_id,
                                'combinations_json': json.loads(json.dumps(combos))
                            }, on_conflict='website_id').execute()
                            st.success(f"成功保存了 '{selected_website}' 的组合策略！")
                            st.cache_data.clear()
                        except Exception as e:
                            st.error(f"保存失败: {e}")

            st.markdown("---")
            st.subheader("总数校验")
            if total_combo_count == count: st.success(f"✅ **校验成功!** 已分配链接数 ({total_combo_count}) 与所需总数 ({count}) 相符。")
            else: st.warning(f"⚠️ **规划未完成。** 当前已分配 {total_combo_count} / {count} 个链接。")

        with right_col:
            st.subheader("📖 锚文本参考")
            with st.expander("➕ 添加新的锚文本", expanded=True):
                with st.form(key=f"add_anchor_form_{selected_website}", clear_on_submit=True):
                    PROJECT_MAP = {"ocm": 1, "mkp": 2}
                    new_anchor_text = st.text_input("新锚文本内容", placeholder="例如：最好的SEO工具")
                    sv_col, project_type_col = st.columns(2)
                    with sv_col: new_anchor_sv = st.number_input("搜索量 (可选)", min_value=0, value=0)
                    with project_type_col: new_project_type = st.radio("项目", options=['ocm', 'mkp'], index=1, horizontal=True)
                    submit_button = st.form_submit_button(label="✅ 确认添加")
                    if submit_button and new_anchor_text:
                        selected_project_id = PROJECT_MAP.get(new_project_type)
                        is_duplicate = not anchors_df[(anchors_df['project_id'] == selected_project_id) & (anchors_df['text'] == new_anchor_text)].empty
                        if is_duplicate:
                            st.warning(f"锚文本 '{new_anchor_text}' 已经在项目 '{new_project_type}' 中存在。")
                        else:
                            try:
                                data_to_insert = {'project_id': selected_project_id, 'text': new_anchor_text, 'search_volume': new_anchor_sv}
                                response = supabase.table('anchor_texts').insert(data_to_insert).execute()
                                if response.data:
                                    st.success(f"成功添加锚文本: '{new_anchor_text}'")
                                    st.cache_data.clear()
                                    st.rerun()
                                else:
                                    st.error("操作失败：数据库未返回成功信息。请检查RLS策略。")
                            except Exception as e:
                                st.error(f"数据库插入时发生错误: {e}")
                                
            search_term = st.text_input("搜索锚文本:", placeholder="输入关键词过滤...", key=f"search_{selected_website}")
            display_anchors_df = project_anchors_df[['text', 'search_volume']].sort_values('search_volume', ascending=False).reset_index(drop=True)
            if search_term: display_anchors_df = display_anchors_df[display_anchors_df['text'].str.contains(search_term, case=False, na=False)]
            st.dataframe(display_anchors_df, width='stretch')

    st.markdown("---")
    st.header("📊 3. 执行分配与预览")
    if st.button("🚀 执行分配并生成预览", type="primary", width='stretch'):
        all_previews = []
        websites_to_process = [name for name, count in website_counts.items() if sum(c.get('count', 0) for c in st.session_state.combinations.get(name, [])) == count and count > 0]
        if not websites_to_process:
            st.error("没有已完成规划的网站可供预览。")
        else:
            with st.spinner(f"正在为 {len(websites_to_process)} 个网站生成预览..."):
                def random_date(start, end):
                    if start > end: start, end = end, start
                    delta = end - start
                    int_delta = (delta.days * 24 * 60 * 60) + delta.seconds
                    random_second = random.randrange(int_delta) if int_delta > 0 else 0
                    return start + datetime.timedelta(seconds=random_second)
                
                for site_name in websites_to_process:
                    tasks = []
                    combos = st.session_state.combinations[site_name]
                    for combo in combos:
                        if combo['count'] > 0 and combo['a1'] not in ['-- 选择锚文本 --', ''] and combo['u1'] not in ['-- 选择URL --', '']:
                            a2_val = combo.get('a2', '') if combo.get('a2') != '-- 选择锚文本 --' else ''
                            u2_val = combo.get('u2', '') if combo.get('u2') != '-- 选择URL --' else ''
                            for _ in range(combo['count']):
                                tasks.append({'Anchor_1': combo['a1'], 'Url1': combo['u1'], 'Anchor_2': a2_val, 'Url2': u2_val, 'Date': random_date(start_date, end_date).strftime('%Y-%m-%d')})
                    
                    random.shuffle(tasks)
                    site_indices = cart_df[cart_df['client_websites.name'] == site_name].index
                    if len(tasks) == len(site_indices):
                        assigned_df = pd.DataFrame(tasks, index=site_indices)
                        all_previews.append(assigned_df)

            if all_previews:
                master_preview_df = pd.concat(all_previews)
                final_df = cart_df.join(master_preview_df)
                st.session_state.preview_df = final_df
                st.success(f"成功为 {len(websites_to_process)} 个网站分配了链接！")
            else:
                st.error("分配失败，任务数量与链接索引数量不匹配。")

    if st.session_state.preview_df is not None and not st.session_state.preview_df.empty:
        st.markdown("---")
        st.header("✅ 4. 下载订单")
        display_df = st.session_state.preview_df
        final_display_columns = ['Livelink', 'Anchor_1', 'Url1', 'Anchor_2', 'Url2', 'Date', 'Website', 'Vendor', 'Price']
        final_columns_rename = {'vendor_offers.referring_domains.domain_name': 'Livelink', 'client_websites.name': 'Website', 'vendor_offers.vendors.name': 'Vendor', 'vendor_offers.price': 'Price'}
        display_df_renamed = display_df.rename(columns=final_columns_rename)
        final_df_for_display_and_download = display_df_renamed[[col for col in final_display_columns if col in display_df_renamed.columns]]
        st.dataframe(final_df_for_display_and_download, width='stretch')
        
        def convert_df_to_excel(df_to_convert):
            wb = Workbook()
            ws = wb.active
            ws.title = "订单详情"
            for r_idx, row in enumerate(dataframe_to_rows(df_to_convert, index=False, header=True), 1):
                for c_idx, value in enumerate(row, 1):
                    cell = ws.cell(row=r_idx, column=c_idx, value=value)
                    if r_idx == 1:
                        cell.font = Font(bold=True, color="FFFFFF")
                        cell.fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
                        cell.alignment = Alignment(horizontal='center', vertical='center')
                    else:
                        cell.alignment = Alignment(horizontal='left', vertical='center')
            for col in ws.columns:
                max_length = 0
                column = col[0].column_letter
                for cell in col:
                    try:
                        if len(str(cell.value)) > max_length: max_length = len(str(cell.value))
                    except: pass
                adjusted_width = (max_length + 2)
                ws.column_dimensions[column].width = adjusted_width
            virtual_workbook = io.BytesIO()
            wb.save(virtual_workbook)
            return virtual_workbook.getvalue()

        today_str = datetime.date.today().strftime("%Y%m%d")
        all_vendors = final_df_for_display_and_download['Vendor'].dropna().unique().tolist()

        col1, col2 = st.columns([1, 2])
        with col1:
            st.subheader("📥 下载全部")
            st.markdown("合并所有预览数据到一个精美的Excel文件。")
            all_excel_data = convert_df_to_excel(final_df_for_display_and_download)
            st.download_button(
               label="📦 下载全量订单 (Excel)", data=all_excel_data,
               file_name=f"plan_ALL_VENDORS_{today_str}.xlsx",
               mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
               key="download_all"
            )
        with col2:
            st.subheader("📋 按供应商下载")
            st.markdown("为每个选定的供应商生成独立的Excel文件。")
            st.multiselect(
                "选择供应商:", options=all_vendors, 
                key='selected_vendors_for_download',
                label_visibility="collapsed"
            )
            if not st.session_state.selected_vendors_for_download:
                st.info("请从上方选择一个或多个供应商以生成下载链接。")
            else:
                cols = st.columns(4)
                col_idx = 0
                for vendor in st.session_state.selected_vendors_for_download:
                    with cols[col_idx % 4]:
                        vendor_df = final_df_for_display_and_download[final_df_for_display_and_download['Vendor'] == vendor]
                        vendor_excel_data = convert_df_to_excel(vendor_df)
                        file_name = f"plan_for_{vendor.replace(' ', '_')}_{today_str}.xlsx"
                        st.download_button(
                           label=f"⬇️ {vendor} (.xlsx)", data=vendor_excel_data,
                           file_name=file_name,
                           mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                           key=f"download_{vendor}"
                        )
                    col_idx += 1
